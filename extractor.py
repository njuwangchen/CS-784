from openpyxl import load_workbook
import random
import codecs
import re
import xlwt

def predictAndTest(productNamesSet, trueBrand):
    brand_predicted = []
    for productName in productNamesSet:
        list_words_in_productName = re.split("[ ,]+", productName)

        brandFound = False
        if(not brandFound):
            for i in range(len(list_words_in_productName)-2):
                word_two_token = list_words_in_productName[i]+" "+list_words_in_productName[i+1]+" "+list_words_in_productName[i+2]
                if word_two_token.lower() in brand_dict:
                    brand_predicted.append(word_two_token)
                    brandFound = True
                    break
        if(not brandFound):
            for i in range(len(list_words_in_productName)-1):
                word_two_token = list_words_in_productName[i]+" "+list_words_in_productName[i+1]
                if word_two_token.lower() in brand_dict:
                    brand_predicted.append(word_two_token)
                    brandFound = True
                    break

        if(not brandFound):
            for word in list_words_in_productName:
                if word.lower() in brand_dict:
                    brand_predicted.append(word)
                    brandFound = True
                    break

        if(not brandFound):
            brand_predicted.append("None")

    count = 0
    count_TP =0
    count_TN =0
    count_FN=0
    count_FP=0
    count_T_Predicted=0
    count_T_Labeled=0
    for i in range(len(brand_predicted)):
        if brand_predicted[i].lower() != 'none':
            count_T_Predicted = count_T_Predicted+1
        if trueBrand[i].lower() != 'none':
            count_T_Labeled = count_T_Labeled+1
        if(brand_predicted[i].lower() == trueBrand[i].lower()):
            count=count+1
        else:
            print "======" + productNamesSet[i]
            print brand_predicted[i]+","+trueBrand[i]

        if(brand_predicted[i].lower() == trueBrand[i].lower()):
            if(brand_predicted[i].lower() == 'none'):
                count_TN = count_TN+1
            else:
                count_TP = count_TP+1
        else:
            if(brand_predicted[i].lower() == 'none'):
                count_FN = count_FN+1
            else:
                count_FP = count_FP+1
    return (count_TP,count_TN,count_FP,count_FN,count_T_Predicted,count_T_Labeled)

wb = load_workbook('Samples_clean.xlsx')
ws = wb.active

# read the golden data from the excel file
ProductName_sorted=[];
for cellObj in ws.columns[1]:
        ProductName_sorted.append(str(cellObj.value))
BrandName_true_sorted=[];
for cellObj in ws.columns[2]:
        BrandName_true_sorted.append(str(cellObj.value))
# No brand read as "None"

productNames = []
brandNames_true = []

productNames_test = []
brandNames_true_test = []

all_index = range(len(ProductName_sorted))
random.seed(2)
index_for_test = random.sample(all_index, 120)
index_for_train = list(set(all_index) - set(index_for_test))

for i in index_for_test:
    productNames_test.append(ProductName_sorted[i])
    brandNames_true_test.append(BrandName_true_sorted[i])

testWorkbook = xlwt.Workbook()
testSheet = testWorkbook.add_sheet("test")

row = 0
for i in range(len(productNames_test)):
    testSheet.write(row, 0, productNames_test[i])
    testSheet.write(row, 1, brandNames_true_test[i])
    row = row+1

testWorkbook.save("testSet.xls")

for i in index_for_train:
    productNames.append(ProductName_sorted[i])
    brandNames_true.append(BrandName_true_sorted[i])

trainWorkbook = xlwt.Workbook()
trainSheet = trainWorkbook.add_sheet("development")

row = 0
for i in range(len(productNames)):
    trainSheet.write(row, 0, productNames[i])
    trainSheet.write(row, 1, brandNames_true[i])
    row = row+1

trainWorkbook.save("developmentSet.xls")

brand_dict ={}
f = codecs.open('elec_brand_dic.txt', 'r', encoding='utf-8')

for line in f:
        list_line = line.split('\t')
        brand_dict[list_line[0].lower()]=int(list_line[1])
f.close()

correctNumOfTrain = predictAndTest(productNames, brandNames_true)
print correctNumOfTrain

correctNumOfTest = predictAndTest(productNames_test, brandNames_true_test)
print correctNumOfTest